from django.shortcuts import render, redirect
import os
from django.http import FileResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from leaves.models import Leave, LeaveBalance
from accounts.models import User
from django.contrib.auth.hashers import make_password
from datetime import date, timedelta
import calendar as cal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.core.mail import EmailMessage, send_mail
from django.conf import settings
from django.utils import timezone
from django.contrib import messages
from .models import PaySlipRequest
from django.utils import timezone as tz
from datetime import datetime
from django.core.mail import send_mail
from .models import CertificateRequest
from django.core.mail import send_mail
from .models import BackToWork
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from accounts.models import User, Department
from itertools import chain
from django.core.paginator import Paginator
from django.utils import timezone as tz

# ========================================================================================================================================
# Different views for dashboard : 
# class DashboardView(LoginRequiredMixin, View):
#       What each account sees / how it is shown on dashboard (leave count etc...) + Assigns to correct URL/HTML (idem all views)
# class AdjustBalanceView(LoginRequiredMixin, View): HR Adjusts Remaining days
# class CreateAdminUserView(LoginRequiredMixin, View):HR creates HOA, HR or Calendar Access account
# class DeleteOwnAccountView(LoginRequiredMixin, View): HR Deletes his own account when turnover
# class CalendarView(LoginRequiredMixin, View): whole calendar view (what each account sees on their dashboard + color coded)
# class PromoteUserView(LoginRequiredMixin, View): HR Edit User 
# class DeleteUserView(LoginRequiredMixin, View): HR delete User
# class DeleteLeaveView(LoginRequiredMixin, View): HR delete view (days added back to remaining days)
# class UserDetailView(LoginRequiredMixin, View): HR, HOA, HOD, HOS can get an overview of teacher's and admins' dashboard
# class ExportLeavesView(LoginRequiredMixin, View): Excel file 
# class EditUserView(LoginRequiredMixin, View): HR Edit User (name, department,...)
# class ResetBalancesView(LoginRequiredMixin, View): Resets at new year (vacation leaves carried over for Admin employees)
# class RejectLeaveDashboardView(LoginRequiredMixin, View): Leave rejected + Email sent
# class ArchivesView(LoginRequiredMixin, View): HR Achive (NEEDS IT SUPERVISION)
# class DownloadPDFView(LoginRequiredMixin, View): HR can dowload Archives
# class DeleteArchivesView(LoginRequiredMixin, View):
# class DepartmentListView(LoginRequiredMixin, View):
# class DepartmentCreateView(LoginRequiredMixin, View):
# class DepartmentDeleteView(LoginRequiredMixin, View):
# class ChangePasswordView(LoginRequiredMixin, View):
# class BackToWorkView(LoginRequiredMixin, View):
# class PaySlipView(LoginRequiredMixin, View):
# class CertificateView(LoginRequiredMixin, View):
# class HRFileLeaveView(LoginRequiredMixin, View):
#       def _update_balance(self, leave): important allows to define how leaves are deducted from remaining days/ added / different cases 
# class ApproveLeaveDashboardView(LoginRequiredMixin, View):
#      def _update_balance(self, leave): same as HRFileLeave (same logic when HR files auto approved leaves) 
# ========================================================================================================================================

class DashboardView(LoginRequiredMixin, View):
    def get(self, request):
        user = request.user

        if user.role == 'teacher':
            leaves = Leave.objects.filter(user=user)

            sick_used = 0
            special_used = 0
            SICK_TYPES = ['sick_leave']

            balance = LeaveBalance.objects.filter(
                user=user,
                leave_type='vacation_leave'
            ).first()
            total = float(balance.total_days) if balance else 30
            remaining = float(balance.days_remaining) if balance else 30

            for leave in leaves.filter(status='approved'):
                if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                    days = 0.5
                else:
                    days = 0
                    current = leave.start_leave
                    while current <= leave.end_leave:
                        if current.weekday() < 5:
                            days += 1
                        current += timedelta(days=1)

                if leave.leave_type in SICK_TYPES:
                    sick_used += days
                else:
                    special_used += days

            days_used = sick_used + special_used
            real_unpaid = max(0, days_used - total)
            if real_unpaid > 0:
                special_unpaid = min(real_unpaid, special_used)
                sick_unpaid = max(0, real_unpaid - special_unpaid)
            else:
                sick_unpaid = 0
                special_unpaid = 0

            leave_summary = [
                {'type': 'Sick Leave', 'used': sick_used, 'unpaid': sick_unpaid, 'category': 'sick'},
                {'type': 'Special Leave', 'used': special_used, 'unpaid': special_unpaid, 'category': 'special'},
            ]
            total_unpaid = real_unpaid

            return render(request, 'dashboard/employee_dashboard.html', {
                'leaves': leaves,
                'total_days_used': days_used,
                'leave_summary': leave_summary,
                'user_role': user.role,
                'total_unpaid': total_unpaid,
            })

        elif user.role == 'admin':
            leaves = Leave.objects.filter(user=user)
            balances = LeaveBalance.objects.filter(user=user)
            leave_summary = []
            total_days_used = 0

            NO_LIMIT_TYPES = ['maternity_paternity_leave', 'others']

            for balance in balances:
                days_used = 0
                for leave in leaves.filter(status='approved', leave_type=balance.leave_type):
                    if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                        days_used += 0.5
                    else:
                        current = leave.start_leave
                        while current <= leave.end_leave:
                            if current.weekday() < 5:
                                days_used += 1
                            current += timedelta(days=1)

                total_days_used += days_used
                no_limit = balance.total_days == 0 or balance.leave_type in NO_LIMIT_TYPES

                leave_summary.append({
                    'type': balance.leave_type.replace('_', ' ').title(),
                    'used': days_used,
                    'total': None if no_limit else balance.total_days,
                    'remaining': None if no_limit else balance.days_remaining,
                    'carried_over': balance.carried_over,
                    'unpaid': False if no_limit else balance.days_remaining < 0,
                    'category': None,
                })

            return render(request, 'dashboard/employee_dashboard.html', {
                'leaves': leaves,
                'total_days_used': total_days_used,
                'leave_summary': leave_summary,
                'user_role': user.role,
            })

        elif user.role == 'head_of_department':
            dept = user.department
            if dept:
                subordinates = User.objects.filter(department=dept, role='teacher')
            else:
                subordinates = User.objects.filter(superior=user, role='teacher')

            teacher_leaves = Leave.objects.filter(user__in=subordinates)
            own_leaves = Leave.objects.filter(user=user)

            sick_used = 0
            special_used = 0
            SICK_TYPES = ['sick_leave']

            balance = LeaveBalance.objects.filter(
                user=user,
                leave_type='vacation_leave'
            ).first()
            total = float(balance.total_days) if balance else 30
            remaining = float(balance.days_remaining) if balance else 30

            for leave in own_leaves.filter(status='approved'):
                if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                    days = 0.5
                else:
                    days = 0
                    current = leave.start_leave
                    while current <= leave.end_leave:
                        if current.weekday() < 5:
                            days += 1
                        current += timedelta(days=1)

                if leave.leave_type in SICK_TYPES:
                    sick_used += days
                else:
                    special_used += days

            days_used = sick_used + special_used
            real_unpaid = max(0, days_used - total)
            if real_unpaid > 0:
                special_unpaid = min(real_unpaid, special_used)
                sick_unpaid = max(0, real_unpaid - special_unpaid)
            else:
                sick_unpaid = 0
                special_unpaid = 0

            leave_summary = [
                {'type': 'Sick Leave', 'used': sick_used, 'unpaid': sick_unpaid, 'category': 'sick'},
                {'type': 'Special Leave', 'used': special_used, 'unpaid': special_unpaid, 'category': 'special'},
            ]
            total_unpaid = real_unpaid

            return render(request, 'dashboard/hod_dashboard.html', {
                'subordinates': subordinates,
                'leaves': teacher_leaves,
                'own_leaves': own_leaves,
                'leave_summary': leave_summary,
                'total_days_used': days_used,
                'user_role': user.role,
                'total_unpaid': total_unpaid,
            })

        elif user.role == 'head_of_school':
            hods = User.objects.filter(role='head_of_department')
            hod_teams = []
            for hod in hods:
                dept = hod.department
                if dept:
                    hod_teachers = User.objects.filter(department=dept, role='teacher')
                else:
                    hod_teachers = User.objects.filter(superior=hod, role='teacher')
                hod_teams.append({
                    'hod': hod,
                    'teachers': hod_teachers,
                })

            leaves = Leave.objects.filter(
                user__role__in=['teacher', 'head_of_department'],
                status='pending_hos'
            ) | Leave.objects.filter(
                user__role__in=['teacher', 'head_of_department'],
                status__in=['approved', 'rejected']
            )

            return render(request, 'dashboard/hos_dashboard.html', {
                'hod_teams': hod_teams,
                'leaves': leaves,
            })

        elif user.role == 'head_of_admin':
            admins = User.objects.filter(role__in=['admin', 'hr'])
            leaves = Leave.objects.filter(user__in=admins)
            return render(request, 'dashboard/hoa_dashboard.html', {
                'admins': admins,
                'leaves': leaves,
            })

        elif user.role == 'hr':

            all_leaves_qs = Leave.objects.filter(
                user__isnull=False
            ).exclude(user__role='head_of_admin').order_by('-created_at')

            paginator = Paginator(all_leaves_qs, 25)
            page_number = request.GET.get('page')
            all_leaves = paginator.get_page(page_number)

            all_users = User.objects.exclude(role__in=['hr', 'head_of_admin'])
            pending_users = User.objects.filter(is_active=False, is_email_verified=True)

            hods_no_dept = User.objects.filter(role='head_of_department', department=None)
            teachers_no_superior = User.objects.filter(role='teacher', superior=None)
            no_hos = not User.objects.filter(role='head_of_school', is_active=True).exists()
            no_hoa = not User.objects.filter(role='head_of_admin', is_active=True).exists()
            stale_leaves = Leave.objects.filter(
                user__isnull=False,
                status__in=['pending_hod', 'pending_hos', 'pending_hoa'],
                created_at__lt=tz.now() - timedelta(days=7)
            )

            return render(request, 'dashboard/hr_dashboard.html', {
                'all_leaves': all_leaves,
                'all_users': all_users,
                'pending_users': pending_users,
                'today': date.today(),
                'hods_no_dept': hods_no_dept,
                'teachers_no_superior': teachers_no_superior,
                'no_hos': no_hos,
                'no_hoa': no_hoa,
                'stale_leaves': stale_leaves,
            })

        elif user.role == 'scheduling_team':
            return render(request, 'dashboard/scheduling_dashboard.html', {})

        elif user.role == 'calendar_access':
            return render(request, 'dashboard/scheduling_dashboard.html', {
                'user_role': 'calendar_access'
            })

        else:
            return render(request, 'dashboard/employee_dashboard.html', {
                'leaves': [],
                'total_days_used': 0,
            })
        
class AdjustBalanceView(LoginRequiredMixin, View):
    def post(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')

        user_id = request.POST.get('user_id')
        balance_type = request.POST.get('balance_type')
        user = User.objects.get(id=user_id)

        if balance_type == 'teacher':
            days_remaining = float(request.POST.get('days_remaining', 30))
            balance = LeaveBalance.objects.filter(
                user=user, leave_type='vacation_leave'
            ).first()
            if balance :
                balance.days_remaining = days_remaining
                balance.total_days = float(balance.days_used) + days_remaining
                balance.save()
            else:
                LeaveBalance.objects.create(
                    user=user,
                    leave_type='vacation_leave',
                    total_days=days_remaining,
                    days_used=0,
                    days_remaining=days_remaining,
                    carried_over=0,
                )

        elif balance_type == 'admin':
            leave_type = request.POST.get('leave_type')
            days_remaining = float(request.POST.get('days_remaining', 0))
            balance, created = LeaveBalance.objects.get_or_create(
                user=user,
                leave_type=leave_type,
                defaults={'total_days': days_remaining, 'days_used': 0,
                          'days_remaining': days_remaining, 'carried_over': 0}
            )
            if not created:
                balance.days_remaining = days_remaining
                balance.total_days = days_remaining + float(balance.days_used)
                balance.save()

        return redirect('dashboard')

    def _update_balance(self, leave):
        if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
            count = 0.5
        else:
            count = 0
            current = leave.start_leave
            while current <= leave.end_leave:
                if current.weekday() < 5:
                    count += 1
                current += timedelta(days=1)
        if leave.user.role in ['teacher', 'head_of_department']:
            leave_type = 'vacation_leave'
        else:
            leave_type = leave.leave_type

        balance, created = LeaveBalance.objects.get_or_create(
            user=leave.user,
            leave_type=leave_type,
            defaults={'total_days': 30, 'days_used': 0, 'days_remaining': 30, 'carried_over': 0}
        )
        balance.days_used = float(balance.days_used) + count
        balance.days_remaining = float(balance.days_remaining) - count
        balance.save()

        if leave.user.role in ['teacher', 'head_of_department']:
            if balance.days_remaining < 0 and not leave.is_unpaid:
                leave.is_unpaid = True
                leave.save()

class CreateAdminUserView(LoginRequiredMixin, View):
    def get(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')
        return render(request, 'dashboard/create_admin_user.html')

    def post(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        role = request.POST.get('role')
        password = request.POST.get('password')
        User.objects.create(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            role=role,
            password=make_password(password),
            is_active=True,
            is_approved=True,
            is_email_verified=True,
        )
        return redirect('dashboard')


class DeleteOwnAccountView(LoginRequiredMixin, View):
    def post(self, request):
        user = request.user
        other_hr = User.objects.filter(role='hr', is_active=True).exclude(id=user.id)
        if other_hr.count() == 0:
            return HttpResponse("Cannot delete account — you are the only HR. Create another HR first!")
        user.delete()
        return redirect('login')


class CalendarView(LoginRequiredMixin, View):
    def get(self, request):
        user = request.user

        allowed_roles = ['hr', 'head_of_department', 'head_of_school', 'head_of_admin', 'scheduling_team', 'calendar_access']
        if user.role not in allowed_roles:
            return redirect('dashboard')

        month = int(request.GET.get('month', date.today().month))
        year = int(request.GET.get('year', date.today().year))
        filter_type = request.GET.get('type', 'all')
        filter_department = request.GET.get('department', 'all')

        date_start = date(year, month, 1)
        date_end = date(year, month, cal.monthrange(year, month)[1])

        leaves_approved = Leave.objects.filter(
            status='approved',
            start_leave__lte=date_end,
            end_leave__gte=date_start,
        )

        leaves_pending_sick = Leave.objects.filter(
            status__in=['pending_hod', 'pending_hos', 'pending_hoa'],
            leave_type='sick_leave',
            start_leave__lte=date_end,
            end_leave__gte=date_start,
        )

        if user.role == 'head_of_department':
            leaves_approved = leaves_approved.filter(user__superior=user, user__role='teacher')
            leaves_pending_sick = leaves_pending_sick.filter(user__superior=user, user__role='teacher')

        elif user.role == 'head_of_school':
            leaves_approved = leaves_approved.filter(user__role__in=['teacher', 'head_of_department'])
            leaves_pending_sick = leaves_pending_sick.filter(user__role__in=['teacher', 'head_of_department'])
            if filter_department != 'all':
                leaves_approved = leaves_approved.filter(user__department__id=filter_department)
                leaves_pending_sick = leaves_pending_sick.filter(user__department__id=filter_department)

        elif user.role == 'head_of_admin':
            leaves_approved = leaves_approved.filter(user__role__in=['admin', 'hr'])
            leaves_pending_sick = leaves_pending_sick.filter(user__role__in=['admin', 'hr'])

        elif user.role == 'hr':
            if filter_type == 'teacher':
                leaves_approved = leaves_approved.filter(user__role='teacher')
                leaves_pending_sick = leaves_pending_sick.filter(user__role='teacher')
            elif filter_type == 'hod':
                leaves_approved = leaves_approved.filter(user__role='head_of_department')
                leaves_pending_sick = leaves_pending_sick.filter(user__role='head_of_department')
            elif filter_type == 'admin':
                leaves_approved = leaves_approved.filter(user__role__in=['admin', 'hr'])
                leaves_pending_sick = leaves_pending_sick.filter(user__role__in=['admin', 'hr'])
            if filter_department != 'all':
                leaves_approved = leaves_approved.filter(user__department__id=filter_department)
                leaves_pending_sick = leaves_pending_sick.filter(user__department__id=filter_department)

        elif user.role == 'scheduling_team':
            leaves_approved = leaves_approved.filter(user__role__in=['teacher', 'head_of_department'])
            leaves_pending_sick = leaves_pending_sick.filter(user__role__in=['teacher', 'head_of_department'])
            if filter_department != 'all':
                leaves_approved = leaves_approved.filter(user__department__id=filter_department)
                leaves_pending_sick = leaves_pending_sick.filter(user__department__id=filter_department)

        elif user.role == 'calendar_access':
            if filter_type == 'teacher':
                leaves_approved = leaves_approved.filter(user__role='teacher')
                leaves_pending_sick = leaves_pending_sick.filter(user__role='teacher')
            elif filter_type == 'hod':
                leaves_approved = leaves_approved.filter(user__role='head_of_department')
                leaves_pending_sick = leaves_pending_sick.filter(user__role='head_of_department')
            elif filter_type == 'admin':
                leaves_approved = leaves_approved.filter(user__role__in=['admin', 'hr'])
                leaves_pending_sick = leaves_pending_sick.filter(user__role__in=['admin', 'hr'])
            if filter_department != 'all':
                leaves_approved = leaves_approved.filter(user__department__id=filter_department)
                leaves_pending_sick = leaves_pending_sick.filter(user__department__id=filter_department)
        all_leaves = list(chain(leaves_approved, leaves_pending_sick))

        user_totals = {}
        user_balance_totals = {}
        for leave in all_leaves:
            uid = leave.user.id
            if uid not in user_totals:
                user_totals[uid] = 0
                b = LeaveBalance.objects.filter(
                    user=leave.user,
                    leave_type='vacation_leave'
                ).first()
                user_balance_totals[uid] = float(b.total_days) if b else 30
                for l in Leave.objects.filter(user=leave.user, status='approved'):
                    if l.half_day in ['morning', 'afternoon'] and l.start_leave == l.end_leave:
                        user_totals[uid] += 0.5
                    else:
                        cur = l.start_leave
                        while cur <= l.end_leave:
                            if cur.weekday() < 5:
                                user_totals[uid] += 1
                            cur += timedelta(days=1)

        cal_matrix = cal.monthcalendar(year, month)
        calendar_data = []
        for week in cal_matrix:
            week_data = []
            for day_index, day in enumerate(week):
                if day == 0:
                    week_data.append({'day': 0, 'absents': [], 'is_weekend': False})
                else:
                    is_weekend = day_index >= 5
                    current_date = date(year, month, day)
                    absents = []
                    if not is_weekend:
                        for leave in all_leaves:
                            if leave.start_leave <= current_date <= leave.end_leave:
                                if leave.status != 'approved':
                                    color = '#adb5bd'  # grey = pending sick
                                elif user.role == 'scheduling_team':
                                    color = '#1a3a6b'  # scheduling blue paid/unpaid (no distinction)
                                elif leave.is_unpaid:
                                    color = '#dc3545'  # red = unpaid 
                                else:
                                    color = '#1a3a6b'  # blue = paid
                                unpaid_info = ''
                                if (user.role != 'scheduling_team'
                                        and leave.status == 'approved'
                                        and leave.is_unpaid
                                        and leave.user.role in ['teacher', 'head_of_department']):
                                    uid = leave.user.id
                                    total_used = user_totals.get(uid, 0)
                                    balance_total = user_balance_totals.get(uid, 30)
                                    over = max(0, total_used - balance_total)
                                    if over > 0:
                                        unpaid_info = ''

                                absents.append({
                                    'name': (
                                        f"{leave.user.first_name} {leave.user.last_name}"
                                        + (" (AM)" if leave.half_day == 'morning' else " (PM)" if leave.half_day == 'afternoon' else "")
                                        + unpaid_info
                                        + (" (pending)" if leave.status != 'approved' else "")
                                    ),
                                    'color': color,
                                })
                    week_data.append({
                        'day': day,
                        'absents': absents,
                        'is_weekend': is_weekend,
                    })
            calendar_data.append(week_data)
        departments = Department.objects.all()

        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        if month == 12:
            next_month, next_year = 1, year + 1
        else:
            next_month, next_year = month + 1, year

        month_name = cal.month_name[month]

        return render(request, 'dashboard/calendar.html', {
            'calendar_data': calendar_data,
            'month': month,
            'year': year,
            'month_name': month_name,
            'prev_month': prev_month,
            'prev_year': prev_year,
            'next_month': next_month,
            'next_year': next_year,
            'filter_type': filter_type,
            'filter_department': filter_department,
            'departments': departments,
            'user_role': user.role,
        })

class PromoteUserView(LoginRequiredMixin, View):
    def post(self, request, user_id):
        if request.user.role != 'hr':
            return redirect('dashboard')

        target_user = User.objects.get(id=user_id)
        new_role = request.POST.get('new_role')

        allowed_transitions = {
            'teacher': ['head_of_department', 'head_of_school'],
            'head_of_department': ['teacher', 'head_of_school'],
            'head_of_school': ['teacher', 'head_of_department'],
        }

        if target_user.role not in allowed_transitions:
            return HttpResponse("Role change not allowed for this user type.")

        if new_role not in allowed_transitions.get(target_user.role, []):
            return HttpResponse("Invalid role transition.")

        if new_role in ['head_of_department', 'head_of_school']:
            hos = User.objects.filter(role='head_of_school').first()
            target_user.superior = hos
        elif new_role == 'teacher':
            hod = User.objects.filter(role='head_of_department').first()
            target_user.superior = hod

        target_user.role = new_role
        target_user.save()
        return redirect('dashboard')


class DeleteUserView(LoginRequiredMixin, View):
    def post(self, request, user_id):
        if request.user.role != 'hr':
            return redirect('dashboard')

        target_user = User.objects.get(id=user_id)

        if target_user.role == 'head_of_school':
            other_hos = User.objects.filter(role='head_of_school').exclude(id=user_id)
            if not other_hos.exists():
                return HttpResponse("Please reassign a new Head of School before deleting this one.")

        target_user.delete()
        return redirect('dashboard')


class DeleteLeaveView(LoginRequiredMixin, View):
    def post(self, request, leave_id):
        if request.user.role != 'hr':
            return redirect('dashboard')

        leave = Leave.objects.get(id=leave_id)

        # if leave approved : days are added back 
        if leave.status == 'approved':
            user = leave.user

            if user.role in ['teacher', 'head_of_department']:
                balance = LeaveBalance.objects.filter(
                    user=user, leave_type='vacation_leave'
                ).first()
                if balance:
                    if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                        days = 0.5
                    else:
                        days = 0
                        current = leave.start_leave
                        while current <= leave.end_leave:
                            if current.weekday() < 5:
                                days += 1
                            current += timedelta(days=1)
                    balance.days_used = max(0, float(balance.days_used) - days)
                    balance.days_remaining = float(balance.total_days) - float(balance.days_used)
                    balance.save()

            elif user.role == 'admin':
                # admins : categories
                balance = LeaveBalance.objects.filter(
                    user=user, leave_type=leave.leave_type
                ).first()
                if balance:
                    if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                        days = 0.5
                    else:
                        days = 0
                        current = leave.start_leave
                        while current <= leave.end_leave:
                            if current.weekday() < 5:
                                days += 1
                            current += timedelta(days=1)
                    balance.days_used = max(0, float(balance.days_used) - days)
                    balance.days_remaining = float(balance.days_remaining) + days
                    balance.save()

        leave.delete()
        return redirect('dashboard')

class UserDetailView(LoginRequiredMixin, View):
   def get(self, request, user_id):
       allowed = ['hr', 'head_of_department', 'head_of_school', 'head_of_admin']
       if request.user.role not in allowed:
           return redirect('dashboard')

       target_user = User.objects.get(id=user_id)
       leaves = Leave.objects.filter(user=target_user)
       balances = LeaveBalance.objects.filter(user=target_user)

       leave_summary = []
       total_days_used = 0
       total_unpaid = 0
       SICK_TYPES = ['sick_leave']
       NO_LIMIT_TYPES = ['maternity_paternity_leave', 'others']

       if target_user.role in ['teacher', 'head_of_department']:
           sick_used = 0
           special_used = 0

           balance = LeaveBalance.objects.filter(
               user=target_user,
               leave_type='vacation_leave'
           ).first()
           total = float(balance.total_days) if balance else 30
           remaining = float(balance.days_remaining) if balance else 30

           for leave in leaves.filter(status='approved'):
               if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                   days = 0.5
               else:
                   days = 0
                   current = leave.start_leave
                   while current <= leave.end_leave:
                       if current.weekday() < 5:
                           days += 1
                       current += timedelta(days=1)

               if leave.leave_type in SICK_TYPES:
                   sick_used += days
               else:
                   special_used += days

           days_used = sick_used + special_used
           total_days_used = days_used

           real_unpaid = max(0, days_used - total)

           if real_unpaid > 0:
               special_unpaid = min(real_unpaid, special_used)
               sick_unpaid = max(0, real_unpaid - special_unpaid)
           else:
               sick_unpaid = 0
               special_unpaid = 0

           total_unpaid = real_unpaid

           leave_summary = [
               {
                   'type': 'Sick Leave',
                   'used': sick_used,
                   'unpaid': sick_unpaid,
                   'category': 'sick',
                   'remaining': remaining,
                   'total': total,
               },
               {
                   'type': 'Special Leave',
                   'used': special_used,
                   'unpaid': special_unpaid,
                   'category': 'special',
                   'remaining': remaining,
                   'total': total,
               },
           ]

       elif target_user.role == 'admin':
           for balance in balances:
               days_used = 0
               for leave in leaves.filter(status='approved', leave_type=balance.leave_type):
                   if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                       days_used += 0.5
                   else:
                       current = leave.start_leave
                       while current <= leave.end_leave:
                           if current.weekday() < 5:
                               days_used += 1
                           current += timedelta(days=1)

               total_days_used += days_used
               no_limit = balance.total_days == 0 or balance.leave_type in NO_LIMIT_TYPES

               leave_summary.append({
                   'type': balance.leave_type.replace('_', ' ').title(),
                   'used': days_used,
                   'total': None if no_limit else balance.total_days,
                   'remaining': None if no_limit else balance.days_remaining,
                   'carried_over': balance.carried_over,
                   'unpaid': False if no_limit else balance.days_remaining < 0,
                   'category': None,
               })

       elif target_user.role == 'hr':
           
           for leave in leaves.filter(status='approved'):
               if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                   total_days_used += 0.5
               else:
                   current = leave.start_leave
                   while current <= leave.end_leave:
                       if current.weekday() < 5:
                           total_days_used += 1
                       current += timedelta(days=1)


       return render(request, 'dashboard/user_detail.html', {
           'target_user': target_user,
           'leaves': leaves,
           'leave_summary': leave_summary,
           'total_days_used': total_days_used,
           'user_role': target_user.role,
           'total_unpaid': total_unpaid,
       })

    
class ExportLeavesView(LoginRequiredMixin, View):
    def get(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')

        wb = Workbook()
        ws = wb.active
        ws.title = "Leave Requests"

        blue = "1a3a6b"
        light_blue = "e8eef8"

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', start_color=blue, end_color=blue)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_font = Font(name='Arial', bold=True, color=blue, size=14)
        date_font = Font(name='Arial', italic=True, color='666666', size=10)
        cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        alt_fill = PatternFill('solid', start_color=light_blue, end_color=light_blue)
        approved_fill = PatternFill('solid', start_color='d4edda', end_color='d4edda')
        rejected_fill = PatternFill('solid', start_color='f8d7da', end_color='f8d7da')
        pending_fill = PatternFill('solid', start_color='fff3cd', end_color='fff3cd')
        unpaid_fill = PatternFill('solid', start_color='f8d7da', end_color='f8d7da')

        ws.merge_cells('A1:J1')
        ws['A1'] = 'GESM Leave Management — Leave Requests Report'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:J2')
        ws['A2'] = f'Generated on: {date.today().strftime("%d/%m/%Y")}'
        ws['A2'].font = date_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 10

        headers = [
            'Person', 'Role', 'Start Date', 'Last Day',
            'Leave Type', 'Duration', 'Reason',
            'Status', 'Unpaid Days', 'Submitted On'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 40

        leaves = Leave.objects.filter(user__isnull=False).exclude(
            user__role='head_of_admin'
        ).order_by('-created_at')

        for row_idx, leave in enumerate(leaves, 5):
            is_alt = (row_idx % 2 == 0)
            name = f"{leave.user.first_name} {leave.user.last_name}" if leave.user else "Deleted User"
            role = leave.user.role if leave.user else "unknown"

            # HALF DAYS
            if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave.end_leave:
                leave_days = 0.5
                duration = f"0.5 day ({leave.half_day.title()})"
            else:
                leave_days = 0
                cur = leave.start_leave
                while cur <= leave.end_leave:
                    if cur.weekday() < 5:
                        leave_days += 1
                    cur += timedelta(days=1)
                duration = f"{int(leave_days)} day(s)"

            unpaid_days = '—'

            if leave.status == 'approved' and leave.user:

                if leave.is_unpaid:
                    unpaid_days = leave_days

                elif leave.user.role == 'admin':
                    balance = LeaveBalance.objects.filter(
                        user=leave.user,
                        leave_type=leave.leave_type
                    ).first()
                    if balance and balance.days_remaining < 0:
                        unpaid_days = min(leave_days, abs(balance.days_remaining))

            data = [
                name,
                role.replace('_', ' ').title(),
                leave.start_leave.strftime('%d/%m/%Y'),
                leave.end_leave.strftime('%d/%m/%Y'),
                leave.leave_type.replace('_', ' ').title(),
                duration,
                leave.reason_for_leave,
                leave.status.replace('_', ' ').title(),
                unpaid_days,
                timezone.localtime(leave.created_at).strftime('%d/%m/%Y %H:%M'),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.border = thin_border
                cell.alignment = cell_align if col in [1, 7] else center_align

                if col == 8:  # Status
                    if leave.status == 'approved':
                        cell.fill = approved_fill
                        cell.font = Font(name='Arial', size=10, bold=True, color='155724')
                    elif leave.status == 'rejected':
                        cell.fill = rejected_fill
                        cell.font = Font(name='Arial', size=10, bold=True, color='721c24')
                    else:
                        cell.fill = pending_fill
                        cell.font = Font(name='Arial', size=10, bold=True, color='856404')
                elif col == 9 and unpaid_days != '—':  # Unpaid Days
                    cell.fill = unpaid_fill
                    cell.font = Font(name='Arial', size=10, bold=True, color='721c24')
                elif is_alt:
                    cell.fill = alt_fill

            ws.row_dimensions[row_idx].height = 20

        col_widths = [25, 22, 14, 14, 22, 20, 38, 18, 18, 22]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A5'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="GESM_Leave_Report_{date.today().strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response

class EditUserView(LoginRequiredMixin, View):
   def get(self, request, user_id):
       if request.user.role != 'hr':
           return redirect('dashboard')
       target_user = User.objects.get(id=user_id)
       departments = Department.objects.all()
       return render(request, 'dashboard/edit_user.html', {
           'target_user': target_user,
           'departments': departments,
           'roles': [
               ('teacher', 'Teacher'),
               ('admin', 'Admin'),
               ('head_of_department', 'Head of Department'),
               ('head_of_school', 'Head of School'),
               ('head_of_admin', 'Head of Administration'),
               ('hr', 'HR'),
           ]
       })

   def post(self, request, user_id):
       if request.user.role != 'hr':
           return redirect('dashboard')
       target_user = User.objects.get(id=user_id)
       old_role = target_user.role
       old_dept = target_user.department

       target_user.first_name = request.POST.get('first_name')
       target_user.last_name = request.POST.get('last_name')
       target_user.email = request.POST.get('email')
       new_role = request.POST.get('role')
       target_user.role = new_role

       if old_role == 'head_of_department' and new_role != 'head_of_department':
           if old_dept:
               if old_dept.head == target_user:
                   old_dept.head = None
                   old_dept.save()
               for teacher in User.objects.filter(department=old_dept, role='teacher'):
                   teacher.superior = None
                   teacher.save()
           target_user.department = None
       dept_id = request.POST.get('department')
       if dept_id:
           new_dept = Department.objects.get(id=dept_id)

           # retirer le HOD de l'ancien département si changement
           if old_dept and old_dept != new_dept:
               if old_dept.head == target_user:
                   old_dept.head = None
                   old_dept.save()
               if old_role == 'head_of_department':
                   for teacher in User.objects.filter(department=old_dept, role='teacher'):
                       teacher.superior = None
                       teacher.save()

           target_user.department = new_dept

           if new_role == 'head_of_department':
               new_dept.head = target_user
               new_dept.save()
               for teacher in User.objects.filter(department=new_dept, role='teacher'):
                   teacher.superior = target_user
                   teacher.save()

       else:
           if old_role != 'head_of_department' or new_role == 'head_of_department':
               if old_dept:
                   if old_dept.head == target_user:
                       old_dept.head = None
                       old_dept.save()
                   if old_role == 'head_of_department':
                       for teacher in User.objects.filter(department=old_dept, role='teacher'):
                           teacher.superior = None
                           teacher.save()
               target_user.department = None

       if new_role == 'head_of_department':
           hos = User.objects.filter(role='head_of_school').first()
           target_user.superior = hos
       elif new_role == 'teacher':
           dept = target_user.department
           if dept and dept.head:
               target_user.superior = dept.head
           else:
               target_user.superior = None
       elif new_role in ['head_of_school', 'head_of_admin', 'hr', 'admin']:
           target_user.superior = None

       target_user.save()

       if old_role != new_role:
           if new_role in ['teacher', 'head_of_department', 'head_of_school']:
               LeaveBalance.objects.filter(user=target_user).delete()
               LeaveBalance.objects.create(
                   user=target_user,
                   leave_type='vacation_leave',
                   total_days=30, days_used=0,
                   days_remaining=30, carried_over=0,
               )
           elif new_role == 'admin':
               LeaveBalance.objects.filter(user=target_user).delete()
               defaults = {
                   'vacation_leave': 15,
                   'sick_leave': 15,
                   'bereavement_leave': 5,
                   'emergency_leave': 3,
                   'maternity_paternity_leave': 0,
                   'others': 0,
               }
               for leave_type, total in defaults.items():
                   LeaveBalance.objects.create(
                       user=target_user,
                       leave_type=leave_type,
                       total_days=total, days_used=0,
                       days_remaining=total, carried_over=0,
                   )

       return redirect('dashboard')

class ResetBalancesView(LoginRequiredMixin, View):
    def post(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')

        admin_defaults = {
            'vacation_leave': 15,
            'sick_leave': 15,
            'bereavement_leave': 5,
            'emergency_leave': 3,
            'maternity_paternity_leave': 1,
            'others': 1,
        }

        for user in User.objects.filter(role__in=['admin']):
            for leave_type, default_total in admin_defaults.items():
                balance, created = LeaveBalance.objects.get_or_create(
                    user=user,
                    leave_type=leave_type,
                    defaults={
                        'total_days': default_total,
                        'days_used': 0,
                        'days_remaining': default_total,
                        'carried_over': 0,
                    }
                )
                if not created:
                    if leave_type == 'vacation_leave':
                        unused = max(0, balance.days_remaining)
                        new_total = default_total + unused
                        balance.total_days = new_total
                        balance.days_used = 0
                        balance.days_remaining = new_total
                        balance.carried_over = unused
                    else:
                        balance.total_days = default_total
                        balance.days_used = 0
                        balance.days_remaining = default_total
                        balance.carried_over = 0
                    balance.save()
        for user in User.objects.filter(role__in=['teacher', 'head_of_department', 'head_of_school']):
            balance = LeaveBalance.objects.filter(
                user=user, leave_type='vacation_leave'
            ).first()
            if balance:
                balance.total_days = 30
                balance.days_used = 0
                balance.days_remaining = 30
                balance.carried_over = 0
                balance.save()
        Leave.objects.all().delete()
        BackToWork.objects.all().delete()
        PaySlipRequest.objects.all().delete()
        CertificateRequest.objects.all().delete()

        messages.success(request, 'All balances reset and histories cleared for new year.')
        return redirect('dashboard')


class RejectLeaveDashboardView(LoginRequiredMixin, View):
    def post(self, request, leave_id):
        leave = Leave.objects.get(id=leave_id)
        user = leave.user

        allowed_roles = ['head_of_department', 'head_of_school', 'head_of_admin', 'hr']
        if request.user.role not in allowed_roles:
            return redirect('dashboard')

        leave.status = 'rejected'
        leave.is_unpaid = request.POST.get('is_unpaid') == '1'
        leave.rejection_reason = request.POST.get('rejection_reason', '').strip()
        leave.save()

        reason_text = f"\nReason: {leave.rejection_reason}\n" if leave.rejection_reason else ""

        recipients = set()
        recipients.add(user.email)

        if user.role == 'teacher':
            if user.superior:
                recipients.add(user.superior.email)
            for hos in User.objects.filter(role='head_of_school'):
                recipients.add(hos.email)
            for hr in User.objects.filter(role='hr'):
                recipients.add(hr.email)

        elif user.role in ['admin', 'hr']:
            for hoa in User.objects.filter(role='head_of_admin'):
                recipients.add(hoa.email)
            for hr in User.objects.filter(role='hr'):
                recipients.add(hr.email)

        for email_addr in recipients:
            send_mail(
                subject=f'Leave Request Rejected — {user.first_name} {user.last_name}',
                message=(
                    f'Hello,\n\n'
                    f'{user.first_name} {user.last_name} leave request for '
                    f'{leave.leave_type.replace("_", " ").title()} '
                    f'from {leave.start_leave} to {leave.end_leave} '
                    f'has been rejected.\n'
                    f'{reason_text}\n'
                    f'GESM Leave Management'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email_addr],
            )

        return redirect('dashboard')

class ArchivesView(LoginRequiredMixin, View):
    def get(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')

        leaves_with_pdf = Leave.objects.exclude(
            pdf_attachment=''
        ).exclude(
            pdf_attachment=None
        ).order_by('-created_at')
        
        return render(request, 'dashboard/archives.html', {
            'leaves_with_pdf': leaves_with_pdf,
        })


class DownloadPDFView(LoginRequiredMixin, View):
    def get(self, request, leave_id):
        if request.user.role != 'hr':
            return redirect('dashboard')
        
        leave = Leave.objects.get(id=leave_id)
        
        if not leave.pdf_attachment:
            return HttpResponse("No PDF found for this leave request.")
        
        leave.pdf_attachment.seek(0)
        response = FileResponse(
            leave.pdf_attachment,
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{leave.pdf_attachment.name}"'
        return response


class DeleteArchivesView(LoginRequiredMixin, View):
    def post(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')
        
        leaves_with_pdf = Leave.objects.exclude(
            pdf_attachment=''
        ).exclude(
            pdf_attachment=None
        )
        
# ==============================
# Need to ask for IT Supervision
# ==============================
        for leave in leaves_with_pdf:
            if leave.pdf_attachment:
                if os.path.exists(leave.pdf_attachment.path):
                    os.remove(leave.pdf_attachment.path)
                leave.pdf_attachment = None
                leave.save()
        
        return redirect('archives')

class DepartmentListView(LoginRequiredMixin, View):
   def get(self, request):
       if request.user.role != 'hr':
           return redirect('dashboard')
       
       departments = Department.objects.all()
       hods = User.objects.filter(role='head_of_department')

       dept_data = []
       for dept in departments:
           teacher_count = User.objects.filter(
               department=dept, role='teacher'
           ).count()
           if dept.head and dept.head.role != 'head_of_department':
               dept.head = None
               dept.save()
           if not dept.head:
               auto_hod = User.objects.filter(
                   department=dept, role='head_of_department'
               ).first()
               if auto_hod:
                   dept.head = auto_hod
                   dept.save()

           dept_data.append({
               'dept': dept,
               'teacher_count': teacher_count,
           })

       return render(request, 'dashboard/department_list.html', {
           'dept_data': dept_data,
           'hods': hods,
       })

   def post(self, request):
       if request.user.role != 'hr':
           return redirect('dashboard')
       dept_id = request.POST.get('dept_id')
       hod_id = request.POST.get('hod_id')
       dept = Department.objects.get(id=dept_id)
       if hod_id:
           hod = User.objects.get(id=hod_id)
           dept.head = hod
           dept.save()
           for teacher in User.objects.filter(department=dept, role='teacher'):
               teacher.superior = hod
               teacher.save()
       return redirect('department_list')

class DepartmentCreateView(LoginRequiredMixin, View):
    def post(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')
        name = request.POST.get('name')
        if name:
            Department.objects.create(name=name)
        return redirect('department_list')


class DepartmentDeleteView(LoginRequiredMixin, View):
    def post(self, request, dept_id):
        if request.user.role != 'hr':
            return redirect('dashboard')
        dept = Department.objects.get(id=dept_id)
        
        hods_in_dept = User.objects.filter(department=dept, role='head_of_department')
        teachers_in_dept = User.objects.filter(department=dept, role='teacher')
        
        if hods_in_dept.exists() or teachers_in_dept.exists():
            names = ', '.join([f"{u.first_name} {u.last_name}" for u in hods_in_dept])
            messages.warning(request,
                f"Warning: Department '{dept.name}' has been deleted but "
                f"{hods_in_dept.count()} HOD(s) ({names}) and "
                f"{teachers_in_dept.count()} teacher(s) still have it assigned. "
                f"Please reassign them via Edit User."
            )
        
        dept.delete()
        return redirect('department_list')

class ChangePasswordView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'accounts/change_password.html')

    def post(self, request):
        new_username = request.POST.get('new_username', '').strip()
        current_password = request.POST.get('current_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        if new_username and new_username != request.user.username:
            if User.objects.filter(username=new_username).exclude(id=request.user.id).exists():
                messages.error(request, 'This username is already taken.')
                return render(request, 'accounts/change_password.html')
            request.user.username = new_username
            request.user.save()
            messages.success(request, 'Username updated successfully!')

        if current_password or new_password1 or new_password2:
            if not request.user.check_password(current_password):
                messages.error(request, 'Current password is incorrect.')
                return render(request, 'accounts/change_password.html')
            if new_password1 != new_password2:
                messages.error(request, 'New passwords do not match.')
                return render(request, 'accounts/change_password.html')
            if len(new_password1) < 8:
                messages.error(request, 'Password must be at least 8 characters.')
                return render(request, 'accounts/change_password.html')
            request.user.set_password(new_password1)
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, 'Password changed successfully!')

        return redirect('dashboard')
    
class ManageAccountingEmailView(LoginRequiredMixin, View):
    def post(self, request):
        if request.user.role != 'hr': return redirect('dashboard')
        action = request.POST.get('action')
        if action == 'add':
            label = request.POST.get('label')
            email = request.POST.get('email')
            AccountingEmail.objects.get_or_create(email=email, defaults={'label': label})
        elif action == 'delete':
            AccountingEmail.objects.filter(id=request.POST.get('email_id')).delete()
        return redirect('dashboard')

 
class BackToWorkView(LoginRequiredMixin, View):
   ALLOWED_ROLES = ['teacher', 'head_of_department']

   def get(self, request):
       if request.user.role not in self.ALLOWED_ROLES:
           return redirect('dashboard')
       history = BackToWork.objects.filter(user=request.user).order_by('-sent_at')
       return render(request, 'dashboard/back_to_work.html', {'history': history})

   def post(self, request):
       if request.user.role not in self.ALLOWED_ROLES:
           return redirect('dashboard')

       full_name = request.POST.get('full_name', '').strip()
       return_date = request.POST.get('return_date')
       message = request.POST.get('message', '').strip()

       BackToWork.objects.create(
           user=request.user,
           full_name=full_name,
           return_date=return_date,
           message=message,
       )

       recipients = set()

       # Scheduling team
       for sched in User.objects.filter(role='scheduling_team'):
           recipients.add(sched.email)

       # HOS
       for hos in User.objects.filter(role='head_of_school'):
           recipients.add(hos.email)

       # Superieur direct (HOD du prof, ou rien si c'est un HOD)
       if request.user.superior:
           recipients.add(request.user.superior.email)

       for email in recipients:
           send_mail(
               subject=f'Back to Work — {full_name}',
               message=(
                   f'{full_name} will return to work on {return_date}.\n\n'
                   f'{message}\n\nGESM Leave Management'
               ),
               from_email=settings.DEFAULT_FROM_EMAIL,
               recipient_list=[email],
           )

       messages.success(request, 'Back to work notification sent!')
       return redirect('back_to_work')

class PaySlipView(LoginRequiredMixin, View):
    ALLOWED_ROLES = ['teacher', 'admin', 'head_of_department', 'head_of_school', 'hr']
 
    def get(self, request):
        if request.user.role not in self.ALLOWED_ROLES:
            return redirect('dashboard')
        history = PaySlipRequest.objects.filter(user=request.user).order_by('-sent_at')
        return render(request, 'dashboard/pay_slip.html', {'history': history})
 
    def post(self, request):
        if request.user.role not in self.ALLOWED_ROLES:
            return redirect('dashboard')
        full_name = request.POST.get('full_name', '').strip()
        month = request.POST.get('month', '').strip()
        PaySlipRequest.objects.create(
            user=request.user, full_name=full_name, month=month
        )
        send_mail(
            subject=f'Pay Slip Request — {full_name} — {month}',
            message=(
                f'{full_name} has requested a pay slip for {month}.\n\n'
                f'Please process this request.\n\nGESM Leave Management'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=['accounting@gesm.org'],
        )
        messages.success(request, 'Pay slip request sent!')
        return redirect('pay_slip')

class CertificateView(LoginRequiredMixin, View):
    ALLOWED_ROLES = ['teacher', 'admin', 'head_of_department', 'head_of_school']
 
    def get(self, request):
        if request.user.role not in self.ALLOWED_ROLES:
            return redirect('dashboard')
        history = CertificateRequest.objects.filter(user=request.user).order_by('-sent_at')
        return render(request, 'dashboard/certificate.html', {'history': history})
 
    def post(self, request):
        if request.user.role not in self.ALLOWED_ROLES:
            return redirect('dashboard')
        full_name = request.POST.get('full_name', '').strip()
        purpose = request.POST.get('purpose', '').strip()
        with_salary = request.POST.get('with_salary') == 'on'
        CertificateRequest.objects.create(
            user=request.user, full_name=full_name,
            purpose=purpose, with_salary=with_salary
        )
        salary_text = 'with salary details' if with_salary else 'without salary details'
        for hr in User.objects.filter(role='hr'):
            send_mail(
                subject=f'Certificate of Employment Request — {full_name}',
                message=(
                    f'{full_name} has requested a certificate of employment ({salary_text}).\n\n'
                    f'Purpose: {purpose}\n\nGESM Leave Management'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[hr.email],
            )
        messages.success(request, 'Certificate request sent to HR!')
        return redirect('certificate')

class ApproveLeaveDashboardView(LoginRequiredMixin, View):
    def post(self, request, leave_id):
        leave = Leave.objects.get(id=leave_id)
        user = leave.user

        if request.user.role == 'head_of_department' and leave.status == 'pending_hod':
            leave.status = 'pending_hos'
            leave.save()
            for hos in User.objects.filter(role='head_of_school'):
                self._send_email(
                    to=hos.email,
                    subject=f'Leave Request Pending Your Approval — {user.first_name} {user.last_name}',
                    body=(
                        f'Hello,\n\n'
                        f'{user.first_name} {user.last_name} has requested {leave.leave_type.replace("_", " ").title()} leave '
                        f'from {leave.start_leave} to {leave.end_leave}.\n\n'
                        f'Reason: {leave.reason_for_leave}\n\n'
                        f'This request has been approved by the Head of Department '
                        f'and is now pending your approval. Please log in to your dashboard to approve or reject.\n\n'
                        f'GESM Leave Management'
                    ),
                    leave=leave,
                )

        elif request.user.role == 'head_of_school' and leave.status == 'pending_hos':
            leave.status = 'approved'
            leave.is_unpaid = request.POST.get('is_unpaid') == '1'
            leave._original_end = leave.end_leave
            leave.save()
            self._update_balance(leave)

            recipients = set()
            recipients.add(user.email)
            if user.superior:
                recipients.add(user.superior.email)
            for hos in User.objects.filter(role='head_of_school'):
                recipients.add(hos.email)
            for hr in User.objects.filter(role='hr'):
                recipients.add(hr.email)
            for sched in User.objects.filter(role='scheduling_team'):
                recipients.add(sched.email)

            for email_addr in recipients:
                self._send_email(
                    to=email_addr,
                    subject=f'Leave Request Approved — {user.first_name} {user.last_name}',
                    body=(
                        f'Hello,\n\n'
                        f'{user.first_name} {user.last_name} leave request for '
                        f'{leave.leave_type.replace("_", " ").title()} '
                        f'from {leave.start_leave} to {leave.end_leave} '
                        f'has been fully approved.\n\n'
                        f'GESM Leave Management'
                    ),
                )

        elif request.user.role == 'head_of_admin' and leave.status == 'pending_hoa':
            leave.status = 'approved'
            leave.is_unpaid = request.POST.get('is_unpaid') == '1'
            leave._original_end = leave.end_leave
            leave.save()
            self._update_balance(leave)

            recipients = set()
            recipients.add(user.email)
            for hoa in User.objects.filter(role='head_of_admin'):
                recipients.add(hoa.email)
            for hr in User.objects.filter(role='hr'):
                recipients.add(hr.email)

            for email_addr in recipients:
                self._send_email(
                    to=email_addr,
                    subject=f'Leave Request Approved — {user.first_name} {user.last_name}',
                    body=(
                        f'Hello,\n\n'
                        f'{user.first_name} {user.last_name} leave request for '
                        f'{leave.leave_type.replace("_", " ").title()} '
                        f'from {leave.start_leave} to {leave.end_leave} '
                        f'has been approved by the Head of Administration.\n\n'
                        f'GESM Leave Management'
                    ),
                )

        elif request.user.role == 'hr':
            leave.status = 'approved'
            leave._original_end = leave.end_leave
            leave.save()
            self._update_balance(leave)

            self._send_email(
                to=user.email,
                subject=f'Leave Request Approved — {user.first_name} {user.last_name}',
                body=(
                    f'Hello {user.first_name},\n\n'
                    f'Your {leave.leave_type.replace("_", " ").title()} leave request '
                    f'from {leave.start_leave} to {leave.end_leave} '
                    f'has been approved by HR.\n\n'
                    f'GESM Leave Management'
                ),
            )

        return redirect('dashboard')

    def _send_email(self, to, subject, body, leave=None):
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to],
        )
        if leave and leave.pdf_attachment:
            leave.pdf_attachment.seek(0)
            email.attach(
                leave.pdf_attachment.name,
                leave.pdf_attachment.read(),
                'application/pdf'
            )
        email.send()

    def _get_split_date(self, leave, paid_days):
        count = 0
        current = leave.start_leave
        last_paid = leave.start_leave
        while current <= leave._original_end:
            if current.weekday() < 5:
                count += 1
                if count <= int(paid_days):
                    last_paid = current
                else:
                    break
            current += timedelta(days=1)
        return last_paid

    def _create_half_day_split(self, leave, original_end, half_day_date):
        """Cree les leaves AM paid + PM unpaid + suite unpaid pour le cas 0.5j."""

        # Leave AM paid sur le jour du split
        Leave.objects.create(
            user=leave.user,
            leave_type=leave.leave_type,
            start_leave=half_day_date,
            end_leave=half_day_date,
            half_day='morning',
            reason_for_leave=leave.reason_for_leave + ' (paid half day)',
            status='approved',
            is_unpaid=False,
        )

        # Leave PM unpaid sur le meme jour
        Leave.objects.create(
            user=leave.user,
            leave_type=leave.leave_type,
            start_leave=half_day_date,
            end_leave=half_day_date,
            half_day='afternoon',
            reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
            status='approved',
            is_unpaid=True,
        )

        # Leave unpaid pour les jours suivants (si il en reste)
        next_unpaid = half_day_date + timedelta(days=1)
        while next_unpaid.weekday() >= 5:
            next_unpaid += timedelta(days=1)

        if next_unpaid <= original_end:
            Leave.objects.create(
                user=leave.user,
                leave_type=leave.leave_type,
                start_leave=next_unpaid,
                end_leave=original_end,
                half_day='none',
                reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                status='approved',
                is_unpaid=True,
            )

    def _update_balance(self, leave):
        if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave._original_end:
            count = 0.5
        else:
            count = 0
            current = leave.start_leave
            while current <= leave._original_end:
                if current.weekday() < 5:
                    count += 1
                current += timedelta(days=1)

        if leave.user.role in ['teacher', 'head_of_department']:
            leave_type = 'vacation_leave'
        else:
            leave_type = leave.leave_type

        NO_LIMIT_TYPES = ['maternity_paternity_leave', 'others']

        balance, created = LeaveBalance.objects.get_or_create(
            user=leave.user,
            leave_type=leave_type,
            defaults={'total_days': 30, 'days_used': 0, 'days_remaining': 30, 'carried_over': 0}
        )

        if leave.user.role in ['teacher', 'head_of_department']:
            remaining_before = float(balance.days_remaining)

            if count > remaining_before and remaining_before > 0:
                paid_days = remaining_before
                unpaid_days = count - remaining_before
                original_end = leave._original_end

                if paid_days % 1 == 0.5:
                    int_paid = int(paid_days)

                    if int_paid > 0:
                        # Leave 1 : jours entiers paid
                        split_date = self._get_split_date(leave, int_paid)
                        leave.end_leave = split_date
                        leave.is_unpaid = False
                        leave.save()

                        # Prochain jour ouvre = jour AM/PM
                        half_day_date = split_date + timedelta(days=1)
                        while half_day_date.weekday() >= 5:
                            half_day_date += timedelta(days=1)

                        # Leave 2 AM paid + Leave 3 PM unpaid + Leave 4 suite unpaid
                        self._create_half_day_split(leave, original_end, half_day_date)

                    else:
                        # remaining = 0.5 → AM paid + PM unpaid sur le premier jour
                        first_day = leave.start_leave
                        leave.end_leave = first_day
                        leave.half_day = 'morning'
                        leave.is_unpaid = False
                        leave.save()

                        # PM unpaid sur le meme jour
                        Leave.objects.create(
                            user=leave.user,
                            leave_type=leave.leave_type,
                            start_leave=first_day,
                            end_leave=first_day,
                            half_day='afternoon',
                            reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                            status='approved',
                            is_unpaid=True,
                        )

                        # Suite unpaid si il y a d'autres jours
                        next_unpaid = first_day + timedelta(days=1)
                        while next_unpaid.weekday() >= 5:
                            next_unpaid += timedelta(days=1)

                        if next_unpaid <= original_end:
                            Leave.objects.create(
                                user=leave.user,
                                leave_type=leave.leave_type,
                                start_leave=next_unpaid,
                                end_leave=original_end,
                                half_day='none',
                                reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                                status='approved',
                                is_unpaid=True,
                            )

                else:
                    # Split normal : paid_days est un entier
                    split_date = self._get_split_date(leave, int(paid_days))
                    leave.end_leave = split_date
                    leave.is_unpaid = False
                    leave.save()

                    next_day = split_date + timedelta(days=1)
                    while next_day.weekday() >= 5:
                        next_day += timedelta(days=1)

                    Leave.objects.create(
                        user=leave.user,
                        leave_type=leave.leave_type,
                        start_leave=next_day,
                        end_leave=original_end,
                        half_day='none',
                        reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                        status='approved',
                        is_unpaid=True,
                    )

                balance.days_used = float(balance.days_used) + paid_days
                balance.days_remaining = 0
                balance.save()

                balance.days_used = float(balance.days_used) + unpaid_days
                balance.days_remaining = float(balance.days_remaining) - unpaid_days
                balance.save()

            elif remaining_before <= 0:
                leave.is_unpaid = True
                leave.save()
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

            else:
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

        elif leave.user.role == 'admin':
            remaining_before = float(balance.days_remaining)
            no_limit = leave_type in NO_LIMIT_TYPES

            if no_limit:
                balance.days_used = float(balance.days_used) + count
                balance.save()
                return

            elif count > remaining_before and remaining_before > 0:
                paid_days = remaining_before
                unpaid_days = count - remaining_before
                original_end = leave._original_end

                if paid_days % 1 == 0.5:
                    int_paid = int(paid_days)

                    if int_paid > 0:
                        split_date = self._get_split_date(leave, int_paid)
                        leave.end_leave = split_date
                        leave.is_unpaid = False
                        leave.save()

                        half_day_date = split_date + timedelta(days=1)
                        while half_day_date.weekday() >= 5:
                            half_day_date += timedelta(days=1)

                        self._create_half_day_split(leave, original_end, half_day_date)

                    else:
                        first_day = leave.start_leave
                        leave.end_leave = first_day
                        leave.half_day = 'morning'
                        leave.is_unpaid = False
                        leave.save()

                        Leave.objects.create(
                            user=leave.user,
                            leave_type=leave.leave_type,
                            start_leave=first_day,
                            end_leave=first_day,
                            half_day='afternoon',
                            reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                            status='approved',
                            is_unpaid=True,
                        )

                        next_unpaid = first_day + timedelta(days=1)
                        while next_unpaid.weekday() >= 5:
                            next_unpaid += timedelta(days=1)

                        if next_unpaid <= original_end:
                            Leave.objects.create(
                                user=leave.user,
                                leave_type=leave.leave_type,
                                start_leave=next_unpaid,
                                end_leave=original_end,
                                half_day='none',
                                reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                                status='approved',
                                is_unpaid=True,
                            )

                else:
                    split_date = self._get_split_date(leave, int(paid_days))
                    leave.end_leave = split_date
                    leave.is_unpaid = False
                    leave.save()

                    next_day = split_date + timedelta(days=1)
                    while next_day.weekday() >= 5:
                        next_day += timedelta(days=1)

                    Leave.objects.create(
                        user=leave.user,
                        leave_type=leave.leave_type,
                        start_leave=next_day,
                        end_leave=original_end,
                        half_day='none',
                        reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                        status='approved',
                        is_unpaid=True,
                    )

                balance.days_used = float(balance.days_used) + paid_days
                balance.days_remaining = 0
                balance.save()

                balance.days_used = float(balance.days_used) + unpaid_days
                balance.days_remaining = float(balance.days_remaining) - unpaid_days
                balance.save()

            elif remaining_before <= 0:
                leave.is_unpaid = True
                leave.save()
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

            else:
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

        else:
            balance.days_used = float(balance.days_used) + count
            balance.days_remaining = float(balance.days_remaining) - count
            balance.save()


class HRFileLeaveView(LoginRequiredMixin, View):

    def get(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')
        all_users = User.objects.filter(
            role__in=['teacher', 'head_of_department', 'admin', 'hr'],
            is_active=True
        ).order_by('role', 'last_name')
        return render(request, 'dashboard/hr_file_leave.html', {
            'all_users': all_users,
            'leave_types': Leave.LEAVE_CHOICES,
        })

    def post(self, request):
        if request.user.role != 'hr':
            return redirect('dashboard')

        user_id = request.POST.get('user_id')
        leave_type = request.POST.get('leave_type')
        half_day = request.POST.get('half_day', 'none')
        reason = request.POST.get('reason_for_leave', 'Filed by HR')
        is_unpaid = request.POST.get('is_unpaid') == 'on'

        start_leave = datetime.strptime(request.POST.get('start_leave'), '%Y-%m-%d').date()
        end_leave = datetime.strptime(request.POST.get('end_leave'), '%Y-%m-%d').date()

        target_user = User.objects.get(id=user_id)

        leave = Leave.objects.create(
            user=target_user,
            leave_type=leave_type,
            start_leave=start_leave,
            end_leave=end_leave,
            half_day=half_day,
            reason_for_leave=reason,
            status='approved',
            is_unpaid=is_unpaid,
        )

        leave._original_end = end_leave
        self._update_balance(leave)

        messages.success(request, f'Leave filed and approved for {target_user.first_name} {target_user.last_name}.')
        return redirect('hr_file_leave')

    def _get_split_date(self, leave, paid_days):
        count = 0
        current = leave.start_leave
        last_paid = leave.start_leave
        while current <= leave._original_end:
            if current.weekday() < 5:
                count += 1
                if count <= int(paid_days):
                    last_paid = current
                else:
                    break
            current += timedelta(days=1)
        return last_paid

    def _create_half_day_split(self, leave, original_end, half_day_date):
        """Cree AM paid + PM unpaid + suite unpaid."""

        Leave.objects.create(
            user=leave.user,
            leave_type=leave.leave_type,
            start_leave=half_day_date,
            end_leave=half_day_date,
            half_day='morning',
            reason_for_leave=leave.reason_for_leave + ' (paid half day)',
            status='approved',
            is_unpaid=False,
        )

        Leave.objects.create(
            user=leave.user,
            leave_type=leave.leave_type,
            start_leave=half_day_date,
            end_leave=half_day_date,
            half_day='afternoon',
            reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
            status='approved',
            is_unpaid=True,
        )

        next_unpaid = half_day_date + timedelta(days=1)
        while next_unpaid.weekday() >= 5:
            next_unpaid += timedelta(days=1)

        if next_unpaid <= original_end:
            Leave.objects.create(
                user=leave.user,
                leave_type=leave.leave_type,
                start_leave=next_unpaid,
                end_leave=original_end,
                half_day='none',
                reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                status='approved',
                is_unpaid=True,
            )

    def _update_balance(self, leave):
        if leave.half_day in ['morning', 'afternoon'] and leave.start_leave == leave._original_end:
            count = 0.5
        else:
            count = 0
            current = leave.start_leave
            while current <= leave._original_end:
                if current.weekday() < 5:
                    count += 1
                current += timedelta(days=1)

        if leave.user.role in ['teacher', 'head_of_department']:
            leave_type = 'vacation_leave'
        else:
            leave_type = leave.leave_type

        NO_LIMIT_TYPES = ['maternity_paternity_leave', 'others']

        balance, created = LeaveBalance.objects.get_or_create(
            user=leave.user,
            leave_type=leave_type,
            defaults={'total_days': 30, 'days_used': 0, 'days_remaining': 30, 'carried_over': 0}
        )

        if leave.user.role in ['teacher', 'head_of_department']:
            remaining_before = float(balance.days_remaining)

            if count > remaining_before and remaining_before > 0:
                paid_days = remaining_before
                unpaid_days = count - remaining_before
                original_end = leave._original_end

                if paid_days % 1 == 0.5:
                    int_paid = int(paid_days)

                    if int_paid > 0:
                        split_date = self._get_split_date(leave, int_paid)
                        leave.end_leave = split_date
                        leave.is_unpaid = False
                        leave.save()

                        half_day_date = split_date + timedelta(days=1)
                        while half_day_date.weekday() >= 5:
                            half_day_date += timedelta(days=1)

                        self._create_half_day_split(leave, original_end, half_day_date)

                    else:
                        first_day = leave.start_leave
                        leave.end_leave = first_day
                        leave.half_day = 'morning'
                        leave.is_unpaid = False
                        leave.save()

                        Leave.objects.create(
                            user=leave.user,
                            leave_type=leave.leave_type,
                            start_leave=first_day,
                            end_leave=first_day,
                            half_day='afternoon',
                            reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                            status='approved',
                            is_unpaid=True,
                        )

                        next_unpaid = first_day + timedelta(days=1)
                        while next_unpaid.weekday() >= 5:
                            next_unpaid += timedelta(days=1)

                        if next_unpaid <= original_end:
                            Leave.objects.create(
                                user=leave.user,
                                leave_type=leave.leave_type,
                                start_leave=next_unpaid,
                                end_leave=original_end,
                                half_day='none',
                                reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                                status='approved',
                                is_unpaid=True,
                            )

                else:
                    split_date = self._get_split_date(leave, int(paid_days))
                    leave.end_leave = split_date
                    leave.is_unpaid = False
                    leave.save()

                    next_day = split_date + timedelta(days=1)
                    while next_day.weekday() >= 5:
                        next_day += timedelta(days=1)

                    Leave.objects.create(
                        user=leave.user,
                        leave_type=leave.leave_type,
                        start_leave=next_day,
                        end_leave=original_end,
                        half_day='none',
                        reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                        status='approved',
                        is_unpaid=True,
                    )

                balance.days_used = float(balance.days_used) + paid_days
                balance.days_remaining = 0
                balance.save()

                balance.days_used = float(balance.days_used) + unpaid_days
                balance.days_remaining = float(balance.days_remaining) - unpaid_days
                balance.save()

            elif remaining_before <= 0:
                leave.is_unpaid = True
                leave.save()
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

            else:
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

        elif leave.user.role == 'admin':
            remaining_before = float(balance.days_remaining)
            no_limit = leave_type in NO_LIMIT_TYPES

            if no_limit:
                balance.days_used = float(balance.days_used) + count
                balance.save()
                return

            elif count > remaining_before and remaining_before > 0:
                paid_days = remaining_before
                unpaid_days = count - remaining_before
                original_end = leave._original_end

                if paid_days % 1 == 0.5:
                    int_paid = int(paid_days)

                    if int_paid > 0:
                        split_date = self._get_split_date(leave, int_paid)
                        leave.end_leave = split_date
                        leave.is_unpaid = False
                        leave.save()

                        half_day_date = split_date + timedelta(days=1)
                        while half_day_date.weekday() >= 5:
                            half_day_date += timedelta(days=1)

                        self._create_half_day_split(leave, original_end, half_day_date)

                    else:
                        first_day = leave.start_leave
                        leave.end_leave = first_day
                        leave.half_day = 'morning'
                        leave.is_unpaid = False
                        leave.save()

                        Leave.objects.create(
                            user=leave.user,
                            leave_type=leave.leave_type,
                            start_leave=first_day,
                            end_leave=first_day,
                            half_day='afternoon',
                            reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                            status='approved',
                            is_unpaid=True,
                        )

                        next_unpaid = first_day + timedelta(days=1)
                        while next_unpaid.weekday() >= 5:
                            next_unpaid += timedelta(days=1)

                        if next_unpaid <= original_end:
                            Leave.objects.create(
                                user=leave.user,
                                leave_type=leave.leave_type,
                                start_leave=next_unpaid,
                                end_leave=original_end,
                                half_day='none',
                                reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                                status='approved',
                                is_unpaid=True,
                            )

                else:
                    split_date = self._get_split_date(leave, int(paid_days))
                    leave.end_leave = split_date
                    leave.is_unpaid = False
                    leave.save()

                    next_day = split_date + timedelta(days=1)
                    while next_day.weekday() >= 5:
                        next_day += timedelta(days=1)

                    Leave.objects.create(
                        user=leave.user,
                        leave_type=leave.leave_type,
                        start_leave=next_day,
                        end_leave=original_end,
                        half_day='none',
                        reason_for_leave=leave.reason_for_leave + ' (unpaid portion)',
                        status='approved',
                        is_unpaid=True,
                    )

                balance.days_used = float(balance.days_used) + paid_days
                balance.days_remaining = 0
                balance.save()

                balance.days_used = float(balance.days_used) + unpaid_days
                balance.days_remaining = float(balance.days_remaining) - unpaid_days
                balance.save()

            elif remaining_before <= 0:
                leave.is_unpaid = True
                leave.save()
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

            else:
                balance.days_used = float(balance.days_used) + count
                balance.days_remaining = float(balance.days_remaining) - count
                balance.save()

        else:
            balance.days_used = float(balance.days_used) + count
            balance.days_remaining = float(balance.days_remaining) - count
            balance.save()